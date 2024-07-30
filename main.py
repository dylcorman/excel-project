# import openpyxl
# from openpyxl import Workbook

# def create_excel_file(file_name):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Data"
    
#     # Add headers
#     ws.append(["Account Number", "Date", "Dollar Amount"])
    
#     wb.save(file_name)
#     print(f"Excel file '{file_name}' created with headers.")

# def add_data_to_excel(file_name, account_number, date, dollar_amount):
#     wb = openpyxl.load_workbook(file_name)
#     ws = wb["Data"]
    
#     # Append data to the sheet
#     ws.append([account_number, date, dollar_amount])
    
#     wb.save(file_name)
#     print(f"Data added to '{file_name}': {account_number}, {date}, {dollar_amount}")

# # Usage example
# file_name = "account_data.xlsx"

# # Create the Excel file with headers (run this once)
# create_excel_file(file_name)

# # Add data to the Excel file
# add_data_to_excel(file_name, "123456", "2024-07-20", 250.75)
# add_data_to_excel(file_name, "789012", "2024-07-21", 145.00)
# add_data_to_excel(file_name, "345678", "2024-07-22", 320.40)


import openpyxl
from openpyxl.styles import PatternFill
import datetime

# File paths
historical_file = '/Users/dylcorman/Downloads/excel-files-main/COPY FOR SEAN - LCVB Occupancy Tax History by Year.xlsx'
new_data_files = [
    '/Users/dylcorman/Downloads/excel-files-main/Occupancy Tax - 10 APR 2024.xlsx',
    '/Users/dylcorman/Downloads/excel-files-main/Occupancy Tax - 11 MAY 2024.xlsx'
]

# Color fill for new data
fill_color = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

def update_historical_data(historical_file, new_data_files):
    # Load the historical data workbook
    wb_hist = openpyxl.load_workbook(historical_file)
    ws_hist = wb_hist.active

    for new_file in new_data_files:
        # Load the new data workbook
        wb_new = openpyxl.load_workbook(new_file)
        ws_new = wb_new.active

        # Iterate through the new data rows
        for row in ws_new.iter_rows(min_row=2, values_only=True):  # Assuming the first row is a header
            account_number, date_str, dollar_amount = row
            date = datetime.datetime.strptime(date_str, "%d %b %Y").date()

            # Find the account number in the historical data
            for hist_row in ws_hist.iter_rows(min_row=2):  # Assuming the first row is a header
                hist_account_number = hist_row[0].value
                if hist_account_number == account_number:
                    # Find the column for the corresponding month
                    for cell in ws_hist[1]:  # Assuming the first row has the month headers
                        if cell.value and isinstance(cell.value, str):
                            try:
                                month_date = datetime.datetime.strptime(cell.value, "%b %Y").date()
                                if month_date.year == date.year and month_date.month == date.month:
                                    # Update the cell with the new amount
                                    ws_hist.cell(row=hist_row[0].row, column=cell.column).value = dollar_amount
                                    # Fill the cell with the color
                                    ws_hist.cell(row=hist_row[0].row, column=cell.column).fill = fill_color
                                    break
                            except ValueError:
                                continue
                    break

    # Save the updated historical data file
    wb_hist.save(historical_file)

# Run the update function
update_historical_data(historical_file, new_data_files)


